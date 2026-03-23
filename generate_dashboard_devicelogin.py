import io, os, json, webbrowser
from datetime import datetime
import requests
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import msal

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ============================================================
#  CONFIGURATION
# ============================================================
SP_SITE        = "roseblanchetn.sharepoint.com"
SP_SITE_PATH   = "/sites/SDAHSESTPA"
SP_BASE_URL    = "https://roseblanchetn.sharepoint.com"
FILE_UNIQUE_ID = "0761FA65-3D84-4B10-B009-8CA5BF050C98"
SHEET_NAME     = "Semoule SSSE"
OUTPUT_EXCEL   = "anomalies_ssse.xlsx"
CLIENT_ID      = "14d82eec-204b-4c2f-b7e8-296a70dab67e"
AUTHORITY      = "https://login.microsoftonline.com/organizations"
SCOPES         = [
    "https://roseblanchetn.sharepoint.com/AllSites.Read",
    "https://roseblanchetn.sharepoint.com/AllSites.Write"
]

COL_DATE   = "Date"
COL_LOT    = "N°lot"
COL_ETAPE  = "Etape"
COL_NOTIF  = "Notif"
COL_ECHANT = "N° de l'échantillon"

MOIS_FR = {1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",
           7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}

# ============================================================
#  REGLES DE DETECTION DES ANOMALIES
#  (col_excel, label_affiche, condition_lambda, cible_texte)
#  Les colonnes vides/NaN sont ignorées automatiquement.
# ============================================================
NUMERIC_CHECKS = [
    ("Humidité (%)",               "Teneur en Eau",   lambda v: v <= 13 or v >= 14.5, "13 < x < 14.5"),
    ("AW",                         "AW",              lambda v: v >= 0.7,              "< 0.7"),
    ("Protéine Brut (%) (+/-0,7)", "Protéine Brute",  lambda v: v <= 10,               "> 10 %"),
    ("Protéine (%)/MS",            "Protéine/MS",     lambda v: v <= 12,               "> 12 %"),
    ("∑ >400µ",                    "G>400µ",          lambda v: v >= 10,               "< 10 %"),
    ("∑ 355;250",                  "G 355-250µ",      lambda v: v <= 40,               "> 40 %"),
    ("∑ < 200µ",                   "G<200µ",          lambda v: v >= 50,               "< 50 %"),
    ("G < 125µ",                   "G<125µ",          lambda v: v >= 10,               "< 10 %"),
    ("Gluten Humide",              "Gluten Humide",   lambda v: v <= 28,               "> 28 %"),
    ("Gluten Index",               "Gluten Index",    lambda v: v <= 65 or v >= 90,    "65 < x < 90"),
    ("Gluten Sec",                 "Gluten Sec",      lambda v: v <= 10,               "> 10 %"),
    ("Col. b",                     "Couleur b",       lambda v: v <= 18,               "> 18"),
    ("Piqûre Noir",                "Piqûre Noir",     lambda v: v >= 10,               "< 10"),
    ("Piqûre Brun",                "Piqûre Brun",     lambda v: v >= 100,              "< 100"),
    ("Cendres (%) (+/- 0,02)",     "Cendres",         lambda v: v >= 1,                "< 1 %"),
    ("T Chute",                    "Temps de Chute",  lambda v: v <= 250,              "> 250"),
]

STRING_CHECKS = [
    ("Embalage (Etanchité,visuel...)", "Emballage",  "C"),
    ("C.Poids",                        "Poids",      "C"),
    ("C .Date",                        "Etiquetage", "C"),
]

# ============================================================
#  ETAPE 1 — Authentification (inchangée)
# ============================================================
def get_token():
    cache = msal.SerializableTokenCache()
    cache_file = ".token_cache.json"
    if os.path.exists(cache_file):
        with open(cache_file) as f:
            cache.deserialize(f.read())
    app = msal.PublicClientApplication(client_id=CLIENT_ID, authority=AUTHORITY, token_cache=cache)
    accounts = app.get_accounts()
    if accounts:
        print(f"  Compte en cache : {accounts[0]['username']}")
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            print("  Token valide trouvé dans le cache")
            _save_cache(cache, cache_file)
            return result["access_token"]
    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        raise Exception(f"Erreur : {flow}")
    print("\n" + "="*55)
    print("  CONNEXION REQUISE")
    print("="*55)
    print(f"\n  1. Ouvre : https://microsoft.com/devicelogin")
    print(f"  2. Entre le code : {flow['user_code']}")
    print(f"  3. Connecte-toi avec ton compte rose-blanche.com")
    print(f"\n  En attente...\n")
    try:
        webbrowser.open("https://microsoft.com/devicelogin")
    except:
        pass
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise Exception(f"Connexion échouée : {result.get('error_description', result)}")
    print("  Connecté avec succès !")
    _save_cache(cache, cache_file)
    return result["access_token"]

def _save_cache(cache, path):
    if cache.has_state_changed:
        with open(path, "w") as f:
            f.write(cache.serialize())

# ============================================================
#  ETAPE 2 — Lecture SharePoint (inchangée)
# ============================================================
def read_excel(token):
    print("  Téléchargement direct via SharePoint...")
    download_url = (
        f"{SP_BASE_URL}/sites/SDAHSESTPA"
        f"/_layouts/15/download.aspx"
        f"?UniqueId={FILE_UNIQUE_ID}"
    )
    headers = {"Authorization": f"Bearer {token}", "Accept": "*/*", "User-Agent": "Mozilla/5.0"}
    resp = requests.get(download_url, headers=headers, timeout=60, allow_redirects=True)
    if resp.status_code != 200 or b"<!DOCTYPE" in resp.content[:200]:
        raise Exception(f"Erreur {resp.status_code} — supprime .token_cache.json et relance")
    df = pd.read_excel(io.BytesIO(resp.content), sheet_name=SHEET_NAME, header=0)
    df.columns = [str(c).strip() for c in df.columns]  # Nettoyage espaces dans noms de colonnes
    print(f"  {len(df)} lignes lues — feuille '{SHEET_NAME}'")
    return df

# ============================================================
#  ETAPE 3 — Détection des anomalies colonne par colonne
#  Chaque anomalie = 1 ligne dans df_anom
#  Les valeurs vides/NaN ne sont jamais comptées comme anomalie
# ============================================================
def _to_float(val):
    """Convertit une valeur en float. Retourne None si vide ou non convertible."""
    if val is None:
        return None
    if isinstance(val, float) and np.isnan(val):
        return None
    try:
        return float(str(val).replace(',', '.').replace(' ', '').strip())
    except (ValueError, TypeError):
        return None

def prepare_data(df: pd.DataFrame):
    df[COL_DATE]  = pd.to_datetime(df[COL_DATE], errors="coerce")
    df[COL_ETAPE] = df[COL_ETAPE].astype(str).str.strip().str.title()
    df["Année"]    = df[COL_DATE].dt.year.astype("Int64").astype(str)
    df["Mois_num"] = df[COL_DATE].dt.month.astype("Int64")

    rows_anom = []

    for _, r in df.iterrows():
        # --- Vérifications numériques ---
        for col, label, cond, cible in NUMERIC_CHECKS:
            v = _to_float(r.get(col))
            if v is not None and cond(v):
                rows_anom.append({
                    "Date"          : r[COL_DATE],
                    "Année"         : r["Année"],
                    "Mois_num"      : r["Mois_num"],
                    "N°lot"         : r[COL_LOT],
                    "N° Echantillon": r[COL_ECHANT],
                    "Etape"         : r[COL_ETAPE],
                    "Parametre"     : label,
                    "Valeur"        : v,
                    "Cible"         : cible,
                    "Notif"         : r.get(COL_NOTIF),
                    "Commentaires"  : r.get("Commentaires"),
                })

        # --- Vérifications texte (Emballage, Poids, Etiquetage) ---
        for col, label, target in STRING_CHECKS:
            v = r.get(col)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                if str(v).strip() != target:
                    rows_anom.append({
                        "Date"          : r[COL_DATE],
                        "Année"         : r["Année"],
                        "Mois_num"      : r["Mois_num"],
                        "N°lot"         : r[COL_LOT],
                        "N° Echantillon": r[COL_ECHANT],
                        "Etape"         : r[COL_ETAPE],
                        "Parametre"     : label,
                        "Valeur"        : str(v).strip(),
                        "Cible"         : target,
                        "Notif"         : r.get(COL_NOTIF),
                        "Commentaires"  : r.get("Commentaires"),
                    })

    df_anom = pd.DataFrame(rows_anom)

    print(f"  Total lignes analysées : {len(df)}")
    print(f"  Total anomalies réelles: {len(df_anom)}")
    if len(df_anom) > 0:
        print(f"  Répartition par paramètre :")
        for param, cnt in df_anom["Parametre"].value_counts().items():
            print(f"    {param:<25} {cnt}")

    return df, df_anom

# ============================================================
#  ETAPE 4 — Génération Excel structuré (4 feuilles)
# ============================================================
def generate_excel(df_all: pd.DataFrame, df_anom: pd.DataFrame) -> str:
    wb = Workbook()

    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )
    PARAM_COLORS = {
        "Piqûre Brun"   : "FDECEA",
        "Piqûre Noir"   : "EDE7F6",
        "G>400µ"        : "FFF8E1",
        "G 355-250µ"    : "FFF8E1",
        "G<200µ"        : "FFF8E1",
        "G<125µ"        : "FFF8E1",
        "Couleur b"     : "E3F2FD",
        "Teneur en Eau" : "E8F5E9",
        "Gluten Humide" : "FFF3E0",
        "Gluten Index"  : "FFF3E0",
        "Gluten Sec"    : "FFF3E0",
        "AW"            : "FDECEA",
        "Cendres"       : "F5F5F5",
        "Temps de Chute": "F5F5F5",
    }

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = HEADER_ALIGN; c.border = BORDER

    def style_row(ws, row, cols, fill=None):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Feuille 1 : Anomalies_Detail ----
    ws1 = wb.active
    ws1.title = "Anomalies_Detail"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 35

    h1 = ["Date", "Année", "Mois", "Semaine", "N° Lot", "N° Échantillon",
          "Étape", "Paramètre", "Valeur mesurée", "Cible", "Notifié", "Commentaires"]
    ws1.append(h1)
    style_header(ws1, 1, len(h1))

    for i, (_, r) in enumerate(df_anom.iterrows(), 2):
        d = r["Date"] if pd.notna(r["Date"]) else None
        ws1.append([
            d,
            str(r["Année"])    if pd.notna(r.get("Année"))    else "",
            int(r["Mois_num"]) if pd.notna(r.get("Mois_num")) else "",
            d.isocalendar()[1] if d else "",
            str(r["N°lot"])         if pd.notna(r.get("N°lot"))         else "",
            str(r["N° Echantillon"])if pd.notna(r.get("N° Echantillon"))else "",
            str(r["Etape"]),
            str(r["Parametre"]),
            r["Valeur"],
            str(r["Cible"]),
            str(r["Notif"]) if pd.notna(r.get("Notif")) else "Non",
            str(r["Commentaires"]) if pd.notna(r.get("Commentaires")) else "",
        ])
        if d:
            ws1.cell(row=i, column=1).number_format = "DD/MM/YYYY"
        fill = PARAM_COLORS.get(str(r["Parametre"]), "FFFFFF")
        style_row(ws1, i, len(h1), fill=fill if i % 2 == 0 else None)

    set_widths(ws1, [13, 8, 7, 9, 10, 16, 16, 20, 15, 16, 10, 30])

    # ---- Feuille 2 : Resume_Mensuel ----
    ws2 = wb.create_sheet("Resume_Mensuel")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 35

    df2 = df_anom.copy()
    df2["Mois_n"] = df2["Date"].dt.month
    df2["An"]     = df2["Date"].dt.year
    grp2 = df2.groupby(["An", "Mois_n"]).agg(
        Nb      =("Parametre", "count"),
        Notifiees=(COL_NOTIF,  lambda x: (x == "Oui").sum()),
        Types   =("Parametre", "nunique")
    ).reset_index().sort_values(["An", "Mois_n"])

    h2 = ["Année", "Mois Num", "Mois", "Nb Anomalies", "Nb Notifiées", "Types Distincts", "Taux Notification (%)"]
    ws2.append(h2)
    style_header(ws2, 1, len(h2))

    for i, (_, r) in enumerate(grp2.iterrows(), 2):
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws2.append([int(r["An"]), int(r["Mois_n"]), MOIS_FR.get(int(r["Mois_n"]), ""),
                    int(r["Nb"]), int(r["Notifiees"]), int(r["Types"]), taux])
        style_row(ws2, i, len(h2), fill="F0F4FA" if i % 2 == 0 else None)
    set_widths(ws2, [10, 10, 14, 16, 16, 16, 22])

    chart2 = BarChart()
    chart2.type = "col"; chart2.title = "Anomalies par mois"
    chart2.add_data(Reference(ws2, min_col=4, min_row=1, max_row=grp2.shape[0]+1), titles_from_data=True)
    chart2.set_categories(Reference(ws2, min_col=3, min_row=2, max_row=grp2.shape[0]+1))
    chart2.width = 20; chart2.height = 12
    ws2.add_chart(chart2, "I2")

    # ---- Feuille 3 : Resume_Parametre ----
    ws3 = wb.create_sheet("Resume_Parametre")
    ws3.row_dimensions[1].height = 35
    total = len(df_anom)

    grp3 = df_anom.groupby("Parametre").agg(
        Nb       =("Parametre", "count"),
        Notifiees=(COL_NOTIF,   lambda x: (x == "Oui").sum()),
        Etapes   =("Etape",     "nunique")
    ).reset_index().sort_values("Nb", ascending=False)

    h3 = ["Paramètre", "Cible", "Nb Anomalies", "% du Total", "Nb Notifiées", "Taux Notification (%)", "Étapes Concernées"]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))

    # Mapping parametre → cible pour affichage
    cible_map = {label: cible for (_, label, _, cible) in NUMERIC_CHECKS}
    cible_map.update({label: target for (_, label, target) in STRING_CHECKS})

    for i, (_, r) in enumerate(grp3.iterrows(), 2):
        pct  = round(r["Nb"] / total * 100, 1)
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws3.append([
            str(r["Parametre"]),
            cible_map.get(str(r["Parametre"]), ""),
            int(r["Nb"]), pct,
            int(r["Notifiees"]), taux,
            int(r["Etapes"])
        ])
        fill = PARAM_COLORS.get(str(r["Parametre"]), "F0F4FA")
        style_row(ws3, i, len(h3), fill=fill)
    set_widths(ws3, [22, 16, 14, 12, 14, 22, 18])

    pie3 = PieChart(); pie3.title = "Répartition par paramètre"
    pie3.add_data(Reference(ws3, min_col=3, min_row=1, max_row=grp3.shape[0]+1), titles_from_data=True)
    pie3.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=grp3.shape[0]+1))
    pie3.width = 18; pie3.height = 12
    ws3.add_chart(pie3, "I2")

    # ---- Feuille 4 : Resume_Etape ----
    ws4 = wb.create_sheet("Resume_Etape")
    ws4.row_dimensions[1].height = 35

    grp4 = df_anom.groupby("Etape").agg(
        Nb       =("Parametre", "count"),
        Notifiees=(COL_NOTIF,   lambda x: (x == "Oui").sum()),
        Types    =("Parametre", "nunique")
    ).reset_index().sort_values("Nb", ascending=False)

    h4 = ["Étape", "Nb Anomalies", "% du Total", "Nb Notifiées", "Taux Notification (%)", "Types Distincts"]
    ws4.append(h4)
    style_header(ws4, 1, len(h4))

    for i, (_, r) in enumerate(grp4.iterrows(), 2):
        pct  = round(r["Nb"] / total * 100, 1)
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws4.append([str(r["Etape"]), int(r["Nb"]), pct, int(r["Notifiees"]), taux, int(r["Types"])])
        style_row(ws4, i, len(h4), fill="F0F4FA" if i % 2 == 0 else None)
    set_widths(ws4, [18, 14, 12, 14, 22, 16])

    chart4 = BarChart(); chart4.type = "bar"; chart4.title = "Anomalies par étape"
    chart4.add_data(Reference(ws4, min_col=2, min_row=1, max_row=grp4.shape[0]+1), titles_from_data=True)
    chart4.set_categories(Reference(ws4, min_col=1, min_row=2, max_row=grp4.shape[0]+1))
    chart4.width = 18; chart4.height = 12
    ws4.add_chart(chart4, "H2")

    wb.save(OUTPUT_EXCEL)
    print(f"  Fichier : {OUTPUT_EXCEL}")
    print(f"  Feuilles : Anomalies_Detail ({len(df_anom)} lignes) | Resume_Mensuel | Resume_Parametre | Resume_Etape")
    return OUTPUT_EXCEL

# ============================================================
#  MAIN
# ============================================================
if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Export Anomalies SSSE → Excel Power BI")
    print("="*55 + "\n")

    print("[1/4] Authentification Microsoft...")
    token = get_token()

    print("\n[2/4] Lecture fichier Excel SharePoint...")
    df_raw = read_excel(token)

    print("\n[3/4] Détection des anomalies colonne par colonne...")
    df_all, df_anom = prepare_data(df_raw)

    print("\n[4/4] Génération du fichier Excel structuré...")
    path = generate_excel(df_all, df_anom)

    import subprocess
    subprocess.run(["git", "add", path])
    subprocess.run(["git", "commit", "-m", f"Update anomalies {datetime.now().strftime('%Y-%m-%d')}"])
    subprocess.run(["git", "push"])

    print("\n" + "="*55)
    print("  TERMINÉ !")
    print("="*55)
    print(f"\n  Total lignes analysées : {len(df_all)}")
    print(f"  Total anomalies réelles: {len(df_anom)}")
    print(f"  Fichier                : {path}")
    print(f"\n  Prochaine étape :")
    print(f"  → Clic droit sur {path} dans Explorer → Download")
    print(f"  → Ouvre dans Power BI Desktop")
    print()